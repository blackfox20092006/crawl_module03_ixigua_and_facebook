import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.firefox.service import Service as FirefoxService
import json
from webdriver_manager.firefox import GeckoDriverManager
from PyQt5.QtWidgets import QApplication,QMainWindow,QMessageBox
from PyQt5 import QtWidgets
import shutil
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QPixmap
from datetime import  timedelta,datetime
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QMessageBox, QPushButton
from selenium.webdriver.chrome.options import Options
from PyQt5.QtWidgets import QApplication,QMainWindow,QMessageBox
from PyQt5.QtWidgets import QFileDialog
from selenium.webdriver.common.action_chains import ActionChains
import os
import datetime
import openpyxl
import yt_dlp
from yt_dlp.utils import download_range_func
import ffmpeg
import os,cv2,sys,requests
from pathlib import Path
from cache.ui import TOOL
import time
#google lens
link_lens = "https://lens.google.com/search?ep=subb&re=df&s=4&p=AbrfA8qYffnZdoMMmfGvrcPyIQ_rWYY4nrTGXrpHAb93VaTYoOCB7L9IXIa_WNIJ3VXaD2jckrd9D0E7srJcEhY4YXDHXM1eowj9I-nxp5O2_rnlvLUraNMLzaQzOSEHWrMzpKOf3k0z3vTbtzhhko5hV7AswTa9HK-lB1ITCjNyf9ny0t15lvDYYHpoufPPL_DwBf4ruTGARRZfyA%3D%3D#lns=W251bGwsbnVsbCxudWxsLG51bGwsbnVsbCxudWxsLG51bGwsIkVrY0tKR1kzTkdZMk1UZGpMVFZrTmpZdE5ERmhNQzFoTjJSaExUSTFZamd6WW1NM01EWXlNeElmUlRSQlYzaFlaMGxxTm1ka09FZHRjM3A0VFdaRGJIVXpNR2h4UzE5Q1p3PT0iXQ=="
profile_path = "C:\\Users\\conta\\AppData\\Local\\Mozilla\\Firefox\\Profiles\\hcowgg31.default-release-1717055333309"
#cookie facebook
cookie = 'ps_n=1;ps_l=1;datr=OOxKZj0GEVhM850KGMAvxSpr;sb=cwBLZnzQDLs9HZzfoEaU7KWT;dpr=0.75;usida=eyJ2ZXIiOjEsImlkIjoiQXNlY2YzYW1vamF4OCIsInRpbWUiOjE3MTcxNDcwMDN9;locale=en_US;c_user=61560251577865;xs=45%3A9Vrmk-OmB7FAtQ%3A2%3A1717170237%3A-1%3A-1;fr=1Mm5lmbgbr2t2rX3K.AWXamstPOP1u3gA4RmnYxDOpNZs.BmWeAQ..AAA.0.0.BmWfA_.AWXdVhCyP-I;presence=C%7B%22t3%22%3A%5B%5D%2C%22utc3%22%3A1717170237708%2C%22v%22%3A1%7D;wd=1366x599;'

from time import sleep
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import requests
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--mute-audio")
chrome_options.add_argument('log-level=3')
chrome_options.add_argument('--disable-dev-shm-usage')
chrome_options.add_argument('--disable-logging')
chrome_options.add_argument('--disable-extensions')
chrome_options.add_argument('--disable-default-apps')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--log-level=3')
# chrome_options.add_argument('--headless')
videos = []
def get_token_user_pass(username, password):
    url = 'https://b-graph.facebook.com/auth/login?email={}&password={}&access_token=6628568379|c1e620fa708a1d5696fb991c1bde5662&method=post'.format(username, password)
    res = requests.get(url).json()
    return res["access_token"]
def crawl_fb(id, token):
    
        #user or pro5
    url = 'https://graph.facebook.com/{}/videos?type=uploaded&access_token={}'.format(str(id), token)
    res = requests.get(url)
    try:
        if res.json()['data'] == []:
            return 'NoData'
        else:
            for i in res.json()['data']:
                title = i["description"]
                time = i["updated_time"]
                try:
                    comments = i["comments"]["data"]  # list
                except:
                    pass
                watch_link = 'https://www.facebook.com/watch/?v=' + i["id"]
                source_link = i["source"]
                nguoi_dang = i["from"]["name"]
                videos.append(
                    {
                        "tieude": title,
                        "nguoidang": nguoi_dang,
                        "thoigian": time,
                        "binhluan": comments,
                        "linkwatch": watch_link,
                        "linktai": source_link
                    }
                )
            return videos
    except:
        return 'Error'
# print(crawl_fb('100080147624102', 'EAABwzLixnjYBO3jZBDP6ZAEWKKPJxkrCiMToN3TMWZAovwduc9ip3yMbc1MmjZAopDEyWkXB6fyzvuBGBkVORzyldzLpzGQIiklZAd0SEQyqx4bNyHJ0IO0rN7PHA6BTTcTT1Ehz4hSWtal8VpkIr91jcy4ptPZBw8ZBttUYcaZCMETd2bGGzvug81Qkb1EvX3Npgs0ZD'))
def crawl_ixigua(id):
    url = 'https://www.ixigua.com/home/{}?list_entrance=homepage&video_card_type=shortvideo'.format(str(id))
    driver = webdriver.Chrome(options=chrome_options)
    driver.get(url)
    end = '已经到底部，没有更多内容了'
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        WebDriverWait(driver, timeout=10)
        if driver.page_source.find(end) != -1:
            break
    html = driver.page_source
    videos = []
    a = BeautifulSoup(html, 'html.parser')
    lista = a.find_all(attrs={'class': True})
    listb = a.find_all('a')
    for i in lista:
        if i in listb:
            try:
                tmp = BeautifulSoup(str(i), 'html.parser')
                title = tmp.find('a')['title']
                href = tmp.find('a')['href']
                if {
                        'title': title,
                        'link': 'https://www.ixigua.com' + href
                    } not in videos:
                    videos.append(
                        {
                            'title': title,
                            'link': 'https://www.ixigua.com' + href
                        }
                    )
            except:
                pass
    return videos

class MAIN():
    def __init__(self):
        import os
        self.tool_ = QMainWindow()
        self.tool__ = TOOL()
        self.tool__.setupUi(self.tool_)
        self.tool__.bd1.clicked.connect(self.run1)
        self.tool__.dow.clicked.connect(self.run3)
        self.tool__.pushButton_2.clicked.connect(self.run4)
        self.tool__.pushButton_3.clicked.connect(self.choose_f)
        self.tool__.bd2.clicked.connect(self.run2)
        self.tool_.show()
    def choose_f(self):
        file_dialog = QFileDialog()
        file_dialog.setWindowTitle("Chọn tệp")
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setViewMode(QFileDialog.Detail)
        file_dialog.setNameFilter("Text files (*.xlsx)")
        if file_dialog.exec():
            selected_files = file_dialog.selectedFiles() 
            if len(selected_files) > 0: 
                selected_file = selected_files[0] 
                self.selected_directory1 = selected_file 
                print(self.selected_directory1) 
                self.tool__.file_c.setText(self.selected_directory1) 
    # HÀM DOWNLOAD VIDEO
    def run3(self): 
        # NẾU NGƯỜI DÙNG TẢI NHANH QUA Ô NHẬP LINK 
        if len(self.tool__.urlvideo1.text()) > 2:    
            url = self.tool__.urlvideo1.text() 
            # REDDIT CÁCH DOWNLOAD RIÊNG NÊN TÁCH RA , CÒN LẠI DOW BẰNG YT DLP OK
            if "reddit" in url:
                            import requests,os
                            try:
                                url = url.split("/?")[0]
                            except:
                                pass
                            from redvid import Downloader
                            import shutil
                            reddit = Downloader()
                            reddit.auto_max = True
                            proxies = {
                            'http': 'http://autoproxy_B3wkVMPd:prt9WAS5qC@5.161.208.60:5000'
                            }
                            reddit.proxies = proxies
                            urlz = url
                            print(urlz)
                            reddit.url = url
                            name_me = reddit.download()
                            name_nf = urlz.split("/")[-1]
                            print(name_nf)
                            shutil.move(name_me,f".\\KQ2\\DOWNLOAD\\{name_nf}.mp4")
            else:
                path_save_mp4 = ".\\KQ2\\DOWNLOAD"
                import os
                path_save_mp4 = os.path.abspath(path_save_mp4) 
                yt_opts = { 'format': 'best[ext=mp4]', 'force_keyframes_at_cuts': True,'outtmpl': f'{path_save_mp4}/%(title)s.%(ext)s' }
                with yt_dlp.YoutubeDL(yt_opts) as ydl:
                    a = ydl.download(url)
                    try:
                        if "AA" in a:
                            print(a.split(" AA")[0])
                            os.remove(a.split(" AA")[0])
                            with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                ydl.download(url)
                    except:
                        with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                ydl.download(url)
        #NẾU NGƯỜI DÙNG NHẬP FILE EXCEL LIST LINK CẦN TẢI
        else:
            workbook = openpyxl.load_workbook(self.selected_directory1)
            worksheet = workbook.active
            data = []
            for row in range(1, worksheet.max_row + 1):
                url = worksheet.cell(row=row, column=1).value

                data.append(url)
            #DUYỆT QUA FILE EXCEL 
            for stt in range(0,len(data)):
                url = data[stt]
                if url != None:
                    if "reddit" in url:
                            import requests,os
                            try:
                                url = url.split("/?")[0]
                            except:
                                pass
                            from redvid import Downloader
                            import shutil
                            reddit = Downloader()
                            reddit.auto_max = True
                            proxies = {
                            'http': 'http://autoproxy_B3wkVMPd:prt9WAS5qC@5.161.208.60:5000'
                            }
                            reddit.proxies = proxies
                            urlz = url
                            print(urlz)
                            reddit.url = url
                            name_me = reddit.download()
                            name_nf = urlz.split("/")[-1]
                            shutil.move(name_me,f".\\KQ2\\DOWNLOAD\\{stt+1} {name_nf}.mp4")
                    else:
                            print(url)
                            path_save_mp4 = ".\\KQ2\\DOWNLOAD"
                            import os
                            path_save_mp4 = os.path.abspath(path_save_mp4) 
                            print(path_save_mp4)
                            now = datetime.datetime.now()
                            hour = now.hour
                            minute = now.minute
                            time_n = f"{hour}_{minute}"
                            yt_opts = { 'format': 'best[ext=mp4]', 'force_keyframes_at_cuts': True,'outtmpl': f'{path_save_mp4}/{f"{stt+1} "}%(title)s.%(ext)s' }
                            with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                a = ydl.download(url)
                                try:
                                    if "AA" in a:
                                        print(a.split(" AA")[0])
                                        os.remove(a.split(" AA")[0])
                                        with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                            ydl.download(url)
                                except:
                                    with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                            ydl.download(url)
    # TẢI VÀ CẮT VIDEO
    def run4(self):
        import csv,os
        if len(self.tool__.time.text()) > 2:
            url = self.tool__.urlvideo2.text()
            path_save_mp4 = ".\\KQ2\\SPLIT"
            path_save_mp4 = os.path.abspath(path_save_mp4) 
            time_ = self.tool__.time.text()
            start_time = int(time_.split(":")[0])
            end_time = int(time_.split(":")[1]) 
            
            if "reddit" in url:
                            import requests,os
                            try:
                                url = url.split("/?")[0]
                            except:
                                pass
                            from redvid import Downloader
                            import shutil
                            reddit = Downloader()
                            proxies = {
                            'http': 'http://autoproxy_B3wkVMPd:prt9WAS5qC@5.161.208.60:5000'
                            }
                            reddit.proxies = proxies
                            reddit.auto_max = True
                            urlz = url
                            print(urlz)
                            reddit.url = url
                            name_me = reddit.download()
                            name_nf = urlz.split("/")[-1]
                            print(name_nf)
                            path_csv = ".\\manifest.csv"
                            path_csv = os.path.abspath(path_csv)
                            with open( path_csv , 'r') as file:
                                reader = csv.reader(file)
                                rows = list(reader)
                            rows[1] = [start_time, end_time, name_nf]
                            with open( path_csv , 'w', newline='') as file:
                                writer = csv.writer(file)
                                writer.writerows(rows)
                            path_s = f".\\KQ2\\SPLIT\\{name_nf}.mp4"
                            path_s = os.path.abspath(path_s)
                            path_runpy = f".\\split_vd.py"
                            path_runpy = os.path.abspath(path_runpy)
                            print(path_runpy)
                            os.system(f"python {path_runpy} -f {path_s} -m {path_csv}")
                            shutil.move(name_me,path_s)
            else:
                yt_opts = {
                    'verbose': True,
                    'download_ranges': download_range_func(None, [(start_time, end_time)]),
                    'force_keyframes_at_cuts': True,'format': 'best[ext=mp4]', 'force_keyframes_at_cuts': True,'outtmpl': f'{path_save_mp4}/%(title)s.%(ext)s' 
                }
                with yt_dlp.YoutubeDL(yt_opts) as ydl:
                    a = ydl.download(url)
                    try:
                        if "AA" in a:
                            print(a.split(" AA")[0])
                            os.remove(a.split(" AA")[0])
                            with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                ydl.download(url)
                    except:
                        with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                ydl.download(url)
        else:
            workbook = openpyxl.load_workbook(self.selected_directory1)
            worksheet = workbook.active
            data = []
            for row in range(1, worksheet.max_row + 1):
                url = worksheet.cell(row=row, column=1).value
                time_begin = worksheet.cell(row=row, column=2).value
                time_end = worksheet.cell(row=row, column=3).value
                data.append([url, time_begin, time_end])
            for stt in range(0,len(data)):
                    info = data[stt]
                    url = info[0]
                    if url != None:
                        import os,csv
                        if "reddit" in url:
                            import requests,os
                            try:
                                url = url.split("/?")[0]
                            except:
                                pass
                            from redvid import Downloader
                            import shutil
                            reddit = Downloader()
                            proxies = {
                            'http': 'http://autoproxy_B3wkVMPd:prt9WAS5qC@5.161.208.60:5000'
                            }
                            reddit.proxies = proxies
                            reddit.auto_max = True
                            urlz = url
                            print(urlz)
                            reddit.url = url
                            name_me = reddit.download()
                            name_nf = urlz.split("/")[-1]
                            print(name_nf)
                            path_csv = ".\\manifest.csv"
                            path_csv = os.path.abspath(path_csv)
                            with open( path_csv , 'r') as file:
                                reader = csv.reader(file)
                                rows = list(reader)
                            rows[1] = [start_time, end_time, name_nf]
                            with open( path_csv , 'w', newline='') as file:
                                writer = csv.writer(file)
                                writer.writerows(rows)
                            path_s = f".\\KQ2\\SPLIT\\{stt+1} {name_nf}.mp4"
                            path_s = os.path.abspath(path_s)
                            path_runpy = f".\\split_vd.py"
                            path_runpy = os.path.abspath(path_runpy)
                            print(path_runpy)
                            os.system(f"python {path_runpy} -f {path_s} -m {path_csv}")
                            shutil.move(name_me,path_s)
                        else:
                                time_bedit = str(info[1]).split(",")
                                print(time_bedit)
                                start_time = int(time_bedit[0])*60*60+int(time_bedit[1])*60+int(time_bedit[2])
                                time_enedit = str(info[2]).split(",")
                                print(time_enedit)
                                end_time = int(time_enedit[0])*60*60+int(time_enedit[1])*60+int(time_enedit[2])
                                print(url)
                                print(start_time)
                                print(end_time)
                                path_save_mp4 = ".\\KQ2\\SPLIT"
                                import os
                                path_save_mp4 = os.path.abspath(path_save_mp4) 
                                now = datetime.datetime.now()
                                hour = now.hour
                                minute = now.minute
                                time_n = f"{hour}_{minute}"
                                yt_opts = {
                                    'verbose': True,
                                    'download_ranges': download_range_func(None, [(start_time, end_time)]),
                                    'force_keyframes_at_cuts': True,'format': 'best[ext=mp4]', 'force_keyframes_at_cuts': True,'outtmpl': f'{path_save_mp4}/{stt+1} %(title)s.%(ext)s' 
                                }
                                with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                    a = ydl.download(url)
                                    try:
                                        if "AA" in a:
                                            print(a.split(" AA")[0])
                                            os.remove(a.split(" AA")[0])
                                            with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                                ydl.download(url)
                                    except:
                                        with yt_dlp.YoutubeDL(yt_opts) as ydl:
                                                ydl.download(url)

    #HÀM NÀY TẢI VIDEO TỪ LINK, SAU ĐÓ TRÍCH CÁC FRAMES RA CÁC ẢNH , LƯU VÀO FOLDER . SAU ĐÓ LẤY ẢNH UP GOOGLE LENS ĐỂ SEARCH KẾT QUẢ LIÊN QUAN
    def run1(self):

        #download video===================================================
        path_save_mp4 = ".\\cache\\mp4"
        path_save_mp4 = os.path.abspath(path_save_mp4) 
        for filename in os.listdir(path_save_mp4):
            self.filename_ = filename
            file_path = os.path.join(path_save_mp4, filename)
            os.remove(file_path)
        yt_opts = { 'format': 'best[ext=mp4]', 'force_keyframes_at_cuts': True,'outtmpl': f'{path_save_mp4}/%(title)s.%(ext)s' }
        ytid = '4cDqaLxrt6Q'
        with yt_dlp.YoutubeDL(yt_opts) as ydl:
            ydl.download(self.tool__.linkvideo.text())
        #extract frame====================================================
        path_video = next(Path(path_save_mp4).glob("*.mp4"), None)
        print(path_video)
        cap = cv2.VideoCapture(f"{path_video}")
        fps = cap.get(cv2.CAP_PROP_FPS)
        total_duration = cap.get(cv2.CAP_PROP_FRAME_COUNT) / fps
        print(total_duration)
        frames = []
        current_time = 0
        step = self.tool__.doubleSpinBox.text()
        step = step.replace(",",".")
        while current_time <= total_duration:
            print(current_time)
            cap.set(cv2.CAP_PROP_POS_MSEC, current_time * 1000) 
            ret, frame = cap.read()
            if ret:
                frames.append(frame)
            current_time += float(step)
        cap.release()
        for i, frame in enumerate(frames):
            cv2.imwrite(f'{path_save_mp4}\\frame_{i+1}.png', frame)
        #crawl link from google lens====================================================
        from selenium import webdriver
        from selenium.webdriver.firefox.service import Service
        from selenium.webdriver.firefox.options import Options

        options = Options()
        options.add_argument("-profile")
        options.add_argument(profile_path)
        self.driver = webdriver.Firefox(options=options)
        self.driver.get(link_lens)
        main_array = []
        frames = 0
        len_f = len(os.listdir(path_save_mp4))-1
        for filename in os.listdir(path_save_mp4):
            if filename.endswith(".png"):
                frames+=1
                QtWidgets.QApplication.processEvents()
                file_path = os.path.join(path_save_mp4, filename)
                wait = WebDriverWait(self.driver, 4)
                try:
                    wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/header/div[2]/div[2]/div[2]/div/div/div[1]/div/div[1]/div/div[1]/span/div[1]/button"))).click()
                except:
                    self.driver.refresh()
                    time.sleep(2)
                    wait.until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/header/div[2]/div[2]/div[2]/div/div/div[1]/div/div[1]/div/div[1]/span/div[1]/button"))).click()
                time.sleep(1)
                try:
                    WebDriverWait(self.driver, 4).until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'Máy tính')]"))).click()
                except:
                    self.driver.refresh()
                    time.sleep(2)
                    self.driver.find_element(By.XPATH, "/html/body/div[2]/header/div[2]/div[2]/div[2]/div/div/div[1]/div/div[1]/div/div[1]/span/div[1]/button").click()
                    WebDriverWait(self.driver, 4).until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'Máy tính')]"))).click()
                import autoit  
                time.sleep(0.5)
                autoit.send(file_path)
                autoit.send("{ENTER}")
                print(file_path)
                print("============================")
                try:    
                    element = self.driver.find_element(By.XPATH, "/html/body/c-wiz/div/div[2]/div/c-wiz/div/div[2]").get_attribute("outerHTML")
                except:
                    self.driver.refresh()
                    time.sleep(2)
                    element = self.driver.find_element(By.XPATH, "/html/body/c-wiz/div/div[2]/div/c-wiz/div/div[2]").get_attribute("outerHTML")
                import re
                text_array = re.findall(r'href="(.*?)"', element)
                text_array = text_array[0:10]
                main_array.append(text_array)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        col = 2  # Bắt đầu từ cột B
        row = 1
        for sub_array in main_array:
            sheet.cell(row=row, column=col, value=col-1)  # Đánh số cột
            row += 1
            for value in sub_array:
                sheet.cell(row=row, column=col, value=value)
                row += 1
            col += 1
            row = 1
        kq_path = ".\\KQ"
        kq_path = os.path.abspath(kq_path)
        name_ex = os.path.join(kq_path,f"{self.filename_}.xlsx")
        print(name_ex)
        workbook.save(name_ex)
        QtWidgets.QApplication.processEvents()
        print("============================")
    # CRAWL LINK SEARCH CÁC NỀN TẢNG 
    def run2(self):
      
        stop_n = 50
        import requests,json
        keyword = self.tool__.keyword.text()
        self.keywords = keyword
        list_links = []
        begin_date = self.tool__.begin.date()
        year = begin_date.year()
        month = f"{begin_date.month():02d}"
        date_ = f"{begin_date.day():02d}"

        end_date = self.tool__.dateEdit_2.date()
        year1 = end_date.year()
        month1 = f"{end_date.month():02d}"
        date_1 = f"{end_date.day():02d}"       
        data_max = [] 
        type_ = ""
        #TH1: NẾU TRA CỨU BẰNG TỪ KHÓA ( SEARCH )
        if self.tool__.keywords.isChecked():
            #YOUTUBE===========================================
            if self.tool__.ytb.isChecked():
                type_ = "Youtube"
                API_KEY = "AIzaSyCpbyiheuVSmA-aUzaCsnv-dc0pSf9GU9Y" 
                search_url = "https://www.googleapis.com/youtube/v3/search"
                params = {
                    "part" : "snippet",
                    "maxResults" : 50,
                    "q" : keyword,
                    "order":"date",
                    "publishedAfter":f"{year}-{month}-{date_}T00:00:00Z",
                    "publishedBefore":f"{year1}-{month1}-{date_1}T02:00:00Z",
                    "key" : API_KEY
                }
                temp = 0
                while True:
                    try:
                        r = requests.get(search_url, params=params)
                        datas = json.loads(r.text)
                        datas = datas["items"]
                        data_max.append(datas)
                        temp+=len(datas)
                        print(temp)
                    except:
                        print(r.text)
                        break
                for datas in data_max:
                    for data in datas:
                        try:
                            video_id = data["id"]["videoId"]
                            video_id = f"https://www.youtube.com/watch?v={video_id}"
                        except:
                            print(data)
                        list_links.append(video_id)
                print(len(list_links))
            keyword1 = keyword.replace(" ","+")   
            #REDDIT===========================================     
            if self.tool__.reddit.isChecked():
                type_ = "Reddit"
                print("OK")
                url = f"https://www.reddit.com/search/?q={keyword1}&type=media"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                options.binary_location = "C:\\Users\\conta\AppData\\Local\\Mozilla\\Firefox" #CP
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='0.3'")
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                import re
                import time

                max_scroll_loop_count = 10
                scroll_loop_count = 0
                x = 0
                y = 0
                while True:
                        
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        element = self.driver.find_element(By.XPATH, "/html/body/shreddit-app/dsa-transparency-modal-provider/search-dynamic-id-cache-controller/div/div/div[1]/div[2]/main/div/search-media-feed").get_attribute("outerHTML")
                        text_array = re.findall(r'href="(.*?)"', element)
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 5:
                            text_array = list(set(text_array))
                            print(text_array)
                            for i, sublist in enumerate(text_array):
                                text_array[i] = ["https://www.reddit.com" + item for item in sublist]
                            list_links = text_array
                            break
            #FACEBOOK===========================================
            if self.tool__.facebook.isChecked():
                type_ = "Facebook"
                list_links = self.crawl_fb()
            #TIK TOK===========================================
            if self.tool__.tiktok.isChecked():
                type_ = "TikTok"
                print("OK")
                url = f"https://www.tiktok.com/search/video?q={keyword}"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='1'")
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                import re
                import time
                x = 0
                y = 0
                while True:
                        
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        time.sleep(1)
                        element = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]/div[1]/div[2]/div/div").get_attribute("outerHTML")
                        text_array = re.findall(r'href="(.*?)"', element)
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 5:
                            text_array = list(set(text_array))
                            text_array = [item for item in text_array if "http" in item]
                            '''for i, sublist in enumerate(text_array):
                                text_array[i] = ["https:" + item for item in sublist]'''
                            print(text_array)
                            print(len(text_array))
                            list_links = text_array
                            break
            #YOUKU==============================================
            if self.tool__.youku.isChecked():
                type_ = "Youku"
                print("OK")
                keyword2 = keyword.replace(" ","%20")
                url = f"https://so.youku.tv/search_video/q_{keyword2}?searchfrom=1"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='1'")
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                import re
                import time
                x = 0
                y = 0
                while True:
                        
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        element = self.driver.find_element(By.XPATH, "/html/body/div/div/div/div[2]/div/div/div[1]").get_attribute("outerHTML")
                        text_array = re.findall(r'href="(.*?)"', element)
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 5:
                            text_array = list(set(text_array))
                            print(text_array)
                            for i in range(0,len(text_array)):
                                text_array[i] = f"https:{text_array[i]}"
                            print(len(text_array))
                            list_links = text_array
                            break
            #IXIGUA==============================================
            if self.tool__.ixigua.isChecked():
                type_ = "Ixigua"
                print("OK")
                keyword2 = keyword.replace(" ","%20")
                url = f"https://www.ixigua.com/search/{keyword2}"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='1'")
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                import re
                import time
                x = 0
                y = 0
                while True:
                        
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        element = self.driver.find_element(By.XPATH, "/html/body/div[1]/div/main/div/div/div[3]/div/div[1]/div").get_attribute("outerHTML")
                        text_array = re.findall(r'href="(.*?)"', element)
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 5:
                            text_array = list(set(text_array))
                            print(text_array)
                            for i in range(0,len(text_array)):
                                text_array[i] = f"https://www.ixigua.com{text_array[i]}"
                            print(len(text_array))
                            list_links = text_array
                            break
            #X==============================================
            if self.tool__.reddit_2.isChecked():
                type_ = "Twiter"
                print("OK")
                url = f"https://x.com/search?q={keyword}&src=typed_query&f=media"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='0.5'")
                
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                import re
                import time
                time.sleep(4)
                x = 0
                y = 0
                while True:
                        
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        element = wait.until(EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div/div/div[2]/main/div/div/div/div/div/div[3]/section/div/div"))).get_attribute("outerHTML")
                        text_array = re.findall(r'href="(.*?)"', element)
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 20:
                            text_array = list(set(text_array))
                            print(text_array)
                            for i in range(0,len(text_array)):
                                text_array[i] = f"https://x.com{text_array[i]}"
                            print(len(text_array))
                            list_links = text_array
                            break
            #douyin==============================================
            if self.tool__.douyin.isChecked():
                type_ = "Douyin"
                print("OK")
                keyword2 = keyword.replace(" ","%20")
                url = f"https://www.douyin.com/search/{keyword2}"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='0.5'")
                import re
                import time
                time.sleep(2)
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                x = 0
                y = 0
                text_array = []
                while True: 
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        try:
                            element = self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/ul").get_attribute("outerHTML")
                        except:
                            self.driver.refresh()
                            time.sleep(2)
                            element = self.driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/div[2]/div/div[1]/div[2]/div[1]/ul").get_attribute("outerHTML")
                        text_array1 = re.findall(r'href="(.*?)"', element)
                        text_array.append(text_array1)
                        unique_text_array = []
                        for sublist in text_array:
                            if sublist not in unique_text_array:
                                unique_text_array.append(sublist)
                        text_array = unique_text_array
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 5:
                            print(text_array)
                            print(len(text_array))
                            list_links = text_array
                            break
        # NẾU TRA CỨU ID USER ( đang cập nhật )
        else:
            link_rs = []
            #FB=========================================== 
            if self.tool__.facebook.isChecked():
                videos = []
                list_rs = crawl_fb(keyword,TOKEN_FB)
            #ixigua=========================================== 
            elif self.tool__.ixigua.isChecked():
                list_rs = crawl_ixigua(keyword)
            #TIK TOK=========================================== 
            elif self.tool__.youtube.isChecked():
                import pandas as pd
                import requests
                import json
                api_key = 'AIzaSyDVYrCFMWgZ8QqOZGFPqN7TtBSg6edrvDA'
                channel_Id = keyword
                url1 = f"https://www.googleapis.com/youtube/v3/channels?part=statistics&key={api_key}&id={channel_Id}"
                channel_info = requests.get(url1)
                json_data1 = json.loads(channel_info.text)
                limit = 10 # how many pages of information you want
                video_Ids = []
                nextPageToken =""
                for i in range(limit):
                    url = f"https://www.googleapis.com/youtube/v3/search?key={api_key}&part=snippet&channelId={channel_Id}&maxResults=20&pageToken={nextPageToken}"
                    data = json.loads(requests.get(url).text)
                for item in data['items']: 
                    video_Id = str(item['id']['videoId'])
                    video_Ids.append(video_Id)  # Storing video Ids for extracting videos information
                nextPageToken = data['nextPageToken']
                data_df = pd.DataFrame(columns=['video_id','channel_id','published_date',
                             'video_title','video_description',
                             'likes','views','comment_count'])
                data_df.head()
                for i,video_Id in enumerate(video_Ids):
                    url = f"https://www.googleapis.com/youtube/v3/videos?part=statistics,snippet&key={api_key}&id={video_Id}"
                    data = json.loads(requests.get(url).text)
                    channel_id = data['items'][0]['snippet']['channelId']      
                    published_date = data['items'][0]['snippet']['publishedAt']    
                    video_title =  data['items'][0]['snippet']['title']     
                    video_description = data['items'][0]['snippet']['description']
                    likes =  data["items"][0]["statistics"]["likeCount"]
                # dislikes = data["items"][0]["statistics"]["dislikeCount"]
                    views = data["items"][0]["statistics"]["viewCount"]
                    comment_count = data["items"][0]["statistics"]['commentCount']
                    info_videos = [video_Id,channel_id,published_date,
                        video_title,video_description,
                        likes,views,comment_count]
                    list_rs.extend(info_videos)
            elif self.tool__.tiktok.isChecked():
                type_ = "TikTok"
                print("OK")
                url = f"https://www.tiktok.com/@{keyword}"
                from selenium import webdriver
                from selenium.webdriver.firefox.service import Service
                from selenium.webdriver.firefox.options import Options
                options = Options()
                options.add_argument("-profile")
                options.add_argument(profile_path)
                self.driver = webdriver.Firefox(options=options)
                self.driver.get(url)
                self.driver.execute_script("document.body.style.zoom='1'")
                max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                import re
                import time
                x = 0
                y = 0
                text_array = []
                while True:
                        
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        wait = WebDriverWait(self.driver, 10)
                        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                        max_scroll_height = self.driver.execute_script("return document.body.scrollHeight")
                        print("LƯỚT")
                        time.sleep(1)
                        element = self.driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div[2]").get_attribute("outerHTML")
                        text_array = re.findall(r'href="(.*?)"', element)
                        if len(text_array) == x:
                            y+=1
                        else:
                            x = len(text_array) 
                            y = 0
                        print(len(text_array))
                        print(x)
                        if y > 10:
                            text_array = list(set(text_array))
                            text_array = [item for item in text_array if f"/video" in item]
                            print(text_array)
                            print(len(text_array))
                            list_links = text_array
                            break
        # LƯU KẾT QUẢ
        path_tt = ".\\KQ"
        path_tt = os.path.abspath(path_tt)
        save_path = os.path.join(path_tt, type_)
        os.makedirs(save_path, exist_ok=True)
        name_ex = os.path.join(save_path, f"{keyword}.xlsx")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        print(list_links)
        for row, text in enumerate(list_links, start=1):
            sheet.cell(row=row, column=1, value=text)
        workbook.save(name_ex)
        print(name_ex)
    
    def crawl_fb(self):
        dtsg = self.get_dtsg(cookie)
        return self.main(cookie, dtsg, self.keywords)
    def get_dtsg(self,cookie):
        headers = {
            'authority': 'www.facebook.com',
            'cookie': cookie
        }
        response = requests.get('https://www.facebook.com/', headers=headers).text
        dtsg = response.split('["DTSGInitData",[],{"token":"')[1].split('",')[0]
        return dtsg
    def main(self,cookie, dtsg, keyword):

        headers = {
            'cookie': cookie
        }
        data = {
            '__aaid': '0',
            '__a': '1',
            'dpr': '1',
            '__ccg': 'EXCELLENT',
            '__comet_req': '15',
            'fb_dtsg': dtsg,
            '__spin_b': 'trunk',
            'fb_api_caller_class': 'RelayModern',
            'fb_api_req_friendly_name': 'SearchCometResultsInitialResultsQuery',
            'variables': '{"count":5,"allow_streaming":false,"args":{"callsite":"COMET_GLOBAL_SEARCH","config":{"exact_match":false,"high_confidence_config":null,"intercept_config":null,"sts_disambiguation":null,"watch_config":null},"context":{"bsid":"","tsid":""},"experience":{"client_defined_experiences":[],"encoded_server_defined_params":null,"fbid":null,"type":"VIDEOS_TAB"},"filters":[],"text":"'+keyword+'"},"cursor":null,"feedbackSource":23,"fetch_filters":true,"renderLocation":"search_results_page","scale":1,"stream_initial_count":0,"useDefaultActor":false,"__relay_internal__pv__CometImmersivePhotoCanUserDisable3DMotionrelayprovider":false,"__relay_internal__pv__IsWorkUserrelayprovider":false,"__relay_internal__pv__IsMergQAPollsrelayprovider":false,"__relay_internal__pv__CometUFIReactionsEnableShortNamerelayprovider":false,"__relay_internal__pv__CometUFIShareActionMigrationrelayprovider":true,"__relay_internal__pv__CometIsAdaptiveUFIEnabledrelayprovider":false,"__relay_internal__pv__StoriesArmadilloReplyEnabledrelayprovider":true,"__relay_internal__pv__StoriesRingrelayprovider":false,"__relay_internal__pv__EventCometCardImage_prefetchEventImagerelayprovider":true}',
            'doc_id': '7686462554808543',
        }
        response = requests.post('https://www.facebook.com/api/graphql/',headers=headers, data=data)
        response= json.loads(response.text)
        print("=======================")
        list_links = []
        for i in range(0,7):
                try:
                    video_id = response["data"]["serpResponse"]["results"]["edges"][i]["relay_rendering_strategy"]["view_model"]["video_metadata_model"]["video"]["id"]
                    print(video_id)
                    video_id = "https://www.facebook.com/" + str(video_id)
                    list_links.append(video_id)
                except:
                    pass
        has_next_page = response['data']['serpResponse']['results']['page_info']['has_next_page']
        end_cursor = response['data']['serpResponse']['results']['page_info']['end_cursor']
        b = 0
        try:
            while has_next_page:
                b+=1
                if b>1000:
                    print(list_links)
                    break
                data = {
                    '__aaid': '0',
                    '__a': '1',
                    'dpr': '1',
                    '__ccg': 'EXCELLENT',
                    '__comet_req': '15',
                    'fb_dtsg': dtsg,
                    '__spin_b': 'trunk',
                    'fb_api_req_friendly_name': 'SearchCometResultsPaginatedResultsQuery',
                    'variables': '{"allow_streaming":false,"args":{"callsite":"COMET_GLOBAL_SEARCH","config":{"exact_match":false,"high_confidence_config":null,"intercept_config":null,"sts_disambiguation":null,"watch_config":null},"context":{"bsid":"16ae6d65-2d7c-4cc5-87c2-3b496dc6b736","tsid":"0.8959232626035896"},"experience":{"client_defined_experiences":[],"encoded_server_defined_params":null,"fbid":null,"type":"VIDEOS_TAB"},"filters":[],"text":"'+keyword+'"},"count":5,"cursor":"'+end_cursor+'","feedLocation":"SEARCH","feedbackSource":23,"fetch_filters":true,"focusCommentID":null,"locale":null,"privacySelectorRenderLocation":"COMET_STREAM","renderLocation":"search_results_page","scale":1,"stream_initial_count":0,"useDefaultActor":false,"__relay_internal__pv__CometImmersivePhotoCanUserDisable3DMotionrelayprovider":false,"__relay_internal__pv__IsWorkUserrelayprovider":false,"__relay_internal__pv__IsMergQAPollsrelayprovider":false,"__relay_internal__pv__CometUFIReactionsEnableShortNamerelayprovider":false,"__relay_internal__pv__CometUFIShareActionMigrationrelayprovider":true,"__relay_internal__pv__CometIsAdaptiveUFIEnabledrelayprovider":false,"__relay_internal__pv__StoriesArmadilloReplyEnabledrelayprovider":true,"__relay_internal__pv__StoriesRingrelayprovider":false,"__relay_internal__pv__EventCometCardImage_prefetchEventImagerelayprovider":true}',
                    'doc_id': '7692157077508905',
                }
                response = requests.post('https://www.facebook.com/api/graphql/',headers=headers, data=data).json()
                print("=======================")
                for i in range(0,7):
                    try:
                        video_id = response["data"]["serpResponse"]["results"]["edges"][i]["relay_rendering_strategy"]["view_model"]["video_metadata_model"]["video"]["id"]
                        print(video_id)
                        video_id = "https://www.facebook.com/" + str(video_id)
                        list_links.append(video_id)
                    except:
                        pass
                has_next_page = response['data']['serpResponse']['results']['page_info']['has_next_page']
                end_cursor = response['data']['serpResponse']['results']['page_info']['end_cursor']
        except:
            print("END")
            print(list_links)
        return list_links
if __name__ == "__main__":
    app = QApplication([])
    ui = MAIN()
    sys.exit(app.exec_())
